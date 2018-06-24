import openpyxl
from datetime import datetime, timedelta
import calendar
import db_utilities as dbu
import excel_config_reader as ecr


class ErrorOccured(RuntimeError):
   def __init__(self, mssg):
      self.Message = mssg

# initiate the list to store the excel read
result_data = []
holiday_data = []
f = open('output.txt','w')

# This function reads the excel file and returns the workbook address to the called function
def readExcelFile():
    try:
        # file name declaration with path
        L_FileName = 'c:\\Temp\\Uniper\\Workplan.xlsx'

        # passing the file name and creating an instance of the workbook with actual values and ignoring the formulas
        wb = openpyxl.load_workbook(L_FileName, data_only='True')
        return (wb)
    except (Exception, FileExistsError) as error:
        print("Error:%s" % error)

def main():
    try:
        wb = readExcelFile()
        # Getting the activesheet for date processing
        # getting the active worksheet
        wrksheet_names = wb.sheetnames
       #read the excel main sheet to get the data
        readMainActivitySheet(wb, wrksheet_names)

        # read the holidays and load it in the holiday_list
        readHolidays(wb, wrksheet_names)

    except (Exception) as error:
        print("Error in main() %s" %error)
    except (ErrorOccured) as e:
        print("Error in main() %s" %e)


#This function reads the excel stream and creates a list with the following:
# Activity name, start date, and end date
# The list is loaded into result_data
def readMainActivitySheet(wkbook,sheetindex):
    try:
        # get the active sheet
        activityName_active_sheet = sheetindex[0]
        # pass the active sheet name
        sheet = wkbook[activityName_active_sheet]

        # get the max count of rows and cols
        m_row = sheet.max_row
        m_col = sheet.max_column
        m_row = m_row + 1

        # Reading the excel sheet from row number 2
        for curr_row in range(2, m_row, 1):
            result_unique_row = []
            for curr_col in range(4, 8, 1):
                if curr_col == 4:
                    # Appending the activity name
                    data = sheet.cell(row=curr_row, column=curr_col)
                    if data.value != None:
                        result_unique_row.append(data.value)
                elif curr_col == 6 or curr_col == 7:
                    # Appending the start date and end date
                    data = sheet.cell(row=curr_row, column=curr_col)
                    # print(data.value)
                    if data.value != None:
                        dtDate = datetime.strptime(data.value, '%B %d, %Y %I:%M %p')
                        result_unique_row.append(dtDate.date())
            result_data.append(result_unique_row)
    except (Exception) as error:
        print ("Error in readMainActivitySheet(): %s" %error)

#This function reads the result_data and does the following:
#  inserts the activity, start date and end date in activities table
# The second part of the function does the following:
#   takes the start date, end date and expands all date in between them
# Function reads the total work days from excel_config.ini file and based on the value of [TotalWorkDays]
# it expands the date. If it is 5, then only dates between monday to friday is generated
# if TotalWorkDays is 6, then dates between monday to saturday is generated.
def expandDates(result_data):
    try:
        if (len(result_data)) == 0:
            raise ErrorOccured("Empty result List")

        # first remove empty list from result_data i.e, if there are empty rows in the excel
        result_data1 = [x for x in result_data if x != []]
        print(result_data1,file=f)

        #get database connection
        db_conn = dbu.getConn()

        # get the total workdays in a week
        tWdays = ecr.getTotalWorkdays()
        planned_hours =8
        totalRecords = len(result_data1)
        #print(totalRecords)
        counter=0
        dDays = []

        for i in range(0,totalRecords):
            for j in range (1,2):
                activityN = result_data1[i][0]
                stDate = result_data1[i][1]
                endDate = result_data1[i][2]
                print(stDate,endDate,file=f)
                # activities table insert
                execSQL = "INSERT INTO ACTIVITIES (ACTIVITY_ID,PLANNED_START,PLANNED_END) VALUES (%s,%s,%s);"
                execData = (activityN,stDate,endDate)
                print(execSQL,execData,file=f)
                dbu.executeQueryWithData(db_conn, execSQL, execData)

                #Now for each activity, expand the dates startDate until end date
                # and insert into the activities_data table
                dd = [stDate + timedelta(days=x) for x in range((endDate - stDate).days + 1)]
                #print(dd)

                for d in dd:
                    execSQL = "INSERT INTO ACTIVITY_DATA (ACTIVITY_ID,DATE,PLANNED_HOURS) VALUES (%s,%s,%s);"
                    # get the weekday
                    wDay = getDayofWeek(d)
                    dstat = checkIfHoliday(d)
                    planned_hours = 8
                    if tWdays == '5': # if its a 5 day work week
                        if dstat == 'w': # if its not a holiday
                            if wDay == 0 or wDay == 1 or wDay == 2 or wDay == 3 or wDay == 4: #monday - friday
                                # activities table insert
                                execData = (activityN, d,planned_hours)
                                print(execSQL, execData,file=f)
                                counter = counter + 1 #comment this line in production
                            elif wDay == 5 or wDay == 6: # if it is a saturday or sunday, insert a NONE for the planned hours
                                planned_hours = None
                                execData = (activityN, d, planned_hours)
                                print(execSQL, execData, file=f)
                                counter = counter + 1  # comment this line in production
                        elif dstat == 'h': # if it is a holiday, insert a NONE for the planned hours
                            planned_hours = None
                            execData = (activityN, d, planned_hours)
                            print(execSQL, execData, file=f)
                            counter = counter + 1  # comment this line in production
                    elif tWdays == '6': # if its a 6 day work week : monday to Saturday
                        if dstat == 'w':
                            if wDay == 0 or wDay == 1 or wDay == 2 or wDay == 3 or wDay == 4 or wDay == 5:
                                execData = (activityN, d, planned_hours)
                                print(execSQL, execData,file=f)
                                counter = counter + 1  #comment this line in production
                            elif wDay == 6: # if it is a saturday or sunday, insert a NONE for the planned hours
                                planned_hours = None
                                execData = (activityN, d, planned_hours)
                                print(execSQL, execData, file=f)
                                counter = counter + 1  # comment this line in production
                        elif dstat == 'h': # if it is a holiday, insert a NONE for the planned hours
                            planned_hours = None
                            execData = (activityN, d, planned_hours)
                            print(execSQL, execData, file=f)
                            counter = counter + 1  # comment this line in production
                print(counter) #comment this line in production
                counter = 0  #comment this line in production

        #db_conn.close()
    except (Exception) as error:
        print("Error in expandDates(): %s" %error)
    except (ErrorOccured) as e:
        print(e.Message)


#check if the given date is a holiday or a working day
def checkIfHoliday(dDate):
    try:
        dayStatus = 'w'
        if dDate in holiday_data:
            #print('I am in holiday')
            dayStatus = 'h'
            return dayStatus
        else:
            return dayStatus
    except (Exception) as error:
        print("Error in checkIfHoliday() %s" %error)


# this function accepts the date as an argument
# and returns the day of the date. Mon =1, tue=2, wed=3 ....sat=6, sun=0
def getDayofWeek(ddDate):
    try:
        strDate = str(ddDate)
        ddDate = datetime.strptime(strDate, "%Y-%m-%d")
        wkday = calendar.weekday(ddDate.year, ddDate.month, ddDate.day)
        return wkday
    except (Exception) as error:
        print("Error in getDayofWeek(): %s" %error)


# This function reads the list of holidays from the excel sheet
# and populates the holiday_list
def readHolidays(wkbook,sheetindex):
    try:
        # get database connection
        db_conn = dbu.getConn()
        stProc = "SELECT holiday from holidays"
        m_row = dbu.executeQueryRes(db_conn, stProc)

        # First make sure that the Holidays sheet exist : Make it the active sheet
        # if Holidays sheet does not exist, then this module is skipped
        #sheet = wkbook['Holidays']
        # get the max count of rows and cols
        #m_row = sheet.max_row
        #m_row = m_row + 1

        numHolidays = len(m_row)
        # Reading the excel sheet from row number 2
        #for curr_row in range(0, numHolidays, 1):
        for row in m_row:
            result_unique_row = []
            # Appending the activity name
            #data = sheet.cell(row=curr_row, column=1)
            if row[0] != None:
               # dtDate = datetime.strptime(row[0], '%B %d,%Y')
               dtDate = row[0].strftime("%Y-%m-%d")
               holiday_data.append(dtDate)
        #print(holiday_data)
    except (Exception) as error:
        print ("Error in readHolidayExcel(): %s" %error)
    except (ErrorOccured) as e:
        print (e)


#---- Main function starts ----
if __name__ == '__main__':
    main()

expandDates(result_data)

'''
Program Logic

main()
1. Calls the function --> readExcelFile()
2. Gets the workbook object
3. Determines the sheet names
4. Calls the function --> readMainActivitySheet(workbook, sheetnames)
5. Call the function --> readHolidaysExcel(wb, wrksheet_names) to populate the Holiday_data list 


readExcelFile()
1. Gets the file name
2. Opens the excel workbook using the wb instance
3. returns the workbook object

readMainActivitySheet(workbook, sheetnames)
1. Reads the first sheet from the workbook - Task_Table
2. Gets the total rows and columns in the sheet
3. Iterate frm 2nd row to end of sheet
3.1 For every row iterate the columns : 4 -> activity name, 6-> start date, 7 -> end date
3.2 Check if the activity name is empty. pushe the activity name to result_unique_row
3.3 check if date is empty. Read the format of the date. The format is important. Even an extra space will 
    crash the program. The format of date eg: Novemeber 1, 2018 8:10 AM (format is very important)
3.4 Push the date into the result_unique_row
3.5 Then push the result_unique_row into result_data list
4. Iterate through rest of the row in the sheet.

expandDates(result_data)
1. Gets the result_data list
2. If the result_data_list is empty, the program will exit out with exception
3. First remove empty lists from the result_data_list
4. Get the database connection string
5. Get the total working days in a week from excel_config.ini file (5 or 6)
6. Iterate through the result_data list and create INSERT string query for Activity_id, start_date, end_date
7. For each activity - > start_date to end_date
7.1 Expand the dates: dd
7.1.1 For each date : check the following:
7.1.2 The day of the week (Monday =0 ... Sunday = 6) --> getDayOfWeek(date)
7.1.3 Next check whether the date is a holiday by calling function --> checkIfHoliday(d)
7.1.3.1 If the date is a holiday, then discard the date
7.1.4 If its a 5 day week, get the dates between 0 to 4
7.1.5 If its a 6 day week, get the dates between 0 to 5
7.1.6 Create a Insert query string with each expanded date and planned hours = 8

readHolidaysExcel(wb, wrksheet_names)
1. Reads the excel with sheet name 'Holidays'
2. If the sheet does not exist, then it just breaks out but the main program continues
3. If the data exists, then populate the holiday_data list 

getDayofWeek(ddDate):
1. This function returns the day of the week 
2. thi function is called from expandDates
'''
